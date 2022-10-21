#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author: Vincent Deng
# @Time: 18/10/2022 3:05 pm
# @Organisation: Veracode

import configparser
import json
import os.path
import sys

import click
from constant import SETTINGS_INIT_DICT, SETTINGS, CREDENTIALS, save_settings
import credentials_commands as credentials
import user_commands as users
import application_commands as applications
import subprocess
import openpyxl

from application_commands import Application, add_applications_to_platform
from user_commands import User, add_users_to_platform
from credentials_commands import activate_credentials


@click.group()
@click.pass_context
def main(ctx):
  """Veracode PoV Automation Tool"""
  if not CREDENTIALS.parent.exists():
    CREDENTIALS.parent.mkdir()
  if not SETTINGS.parent.exists():
    SETTINGS.parent.mkdir()
  if not CREDENTIALS.exists():
    CREDENTIALS.touch()
  if not SETTINGS.exists():
    SETTINGS.touch()
    save_settings(SETTINGS_INIT_DICT)
  with open(SETTINGS, 'r') as fp:
    settings_dict = json.load(fp)
  ctx.ensure_object(dict)
  ctx.obj['setting'] = settings_dict

  config = configparser.ConfigParser()
  config.read(CREDENTIALS)
  ctx.obj['config'] = config

  # in case activated credentials is deleted, fall back to the first credentials
  if len(config.sections()) > 0 \
          and settings_dict['activated_credentials'] not in config.sections():
    settings_dict['activated_credentials'] = config.sections()[0]
    save_settings(settings_dict)


main.add_command(credentials.credentials)
main.add_command(users.users)
main.add_command(applications.applications)


@main.command()
@click.pass_context
def scan(ctx):
  config = ctx.obj['config']
  if not config.sections():
    click.secho(
      'You have not configured Veracode API credentials, '
      'please run \"pov credentials add\" command before running this command.')
    sys.exit(1)
  """Kick Off a Veracode Scan by Java Wrapper"""
  java_wrapper_path = click.prompt(
    'Please paste the absolution path for Java Wrapper').strip()
  app_name = click.prompt(
    'Please paste the Application Name for the scan').strip()
  create_profile = 'true' if click.confirm(
    'Do you want to create this profile (if not exist)?') else 'false'
  file_path = click.prompt(
    'Please paste the full absolute path for the artifact to scan').strip()
  scan_name = click.prompt('Please enter the scan name').strip()

  cmd = ['java', '-jar', java_wrapper_path, '-action',
         'UploadAndScan', '-appname', app_name, '-createprofile',
         create_profile, '-filepath', file_path, '-version', scan_name]
  p = subprocess.Popen(cmd)
  p.communicate()


@main.command('init')
@click.pass_context
@click.option('-e', '--init-excel', prompt=True)
def initialise(ctx, init_excel):
  """Initialise PoV assets including Applications and Users"""
  file_exists = os.path.exists(init_excel)
  if not file_exists:
    click.secho(f'Cannot locate file {init_excel}.')
    sys.exit(1)

  config = ctx.obj['config']
  if not config.sections():
    click.secho(
      'You have not configured Veracode API credentials, '
      'please run \"pov credentials add\" command before running this command.')
    sys.exit(1)
  setting_dict = ctx.obj['setting']
  use_activated_profile = click.confirm(
    f"Your activated credentials is "
    f"\"{setting_dict['activated_credentials']}\""
    f", continue using this profile to run this command?")
  while not use_activated_profile:
    ctx.invoke(activate_credentials)
    use_activated_profile = click.confirm(
      f"Your activated credentials is "
      f"\"{setting_dict['activated_credentials']}\""
      f", continue using this profile to run this command?")

  setting_dict = ctx.obj['setting']
  config = ctx.obj['config']
  workbook = openpyxl.load_workbook('Application_Inventory.xlsx')
  worksheets = workbook.sheetnames

  application_list = []
  user_list = []

  for idx, worksheet_name in enumerate(worksheets):
    worksheet = workbook[worksheet_name]

    if idx == 0:  # this is application sheet
      if worksheet.cell(row=1, column=2).value != 'Application Name':
        click.secho(f'Cannot parse the input excel.', fg='red')
        sys.exit(1)
      for row in range(2, worksheet.max_row):
        application_name = worksheet.cell(row=row, column=2).value
        if application_name is None or application_name == '':
          continue
        application = Application(application_name,
                                  'Veracode Recommended High + SCA', None, '')
        application_list.append(application)

    elif idx == 1:  # this is user sheet
      if worksheet.cell(row=1, column=2).value != 'Email' or \
              worksheet.cell(row=1, column=3).value != 'First Name' or \
              worksheet.cell(row=1, column=4).value != 'Last Name':
        click.secho(f'Cannot parse the input excel.', fg='red')
        sys.exit(1)

      for row in range(2, worksheet.max_row):
        email = worksheet.cell(row=row, column=2).value
        first_name = worksheet.cell(row=row, column=3).value
        last_name = worksheet.cell(row=row, column=4).value

        if email is None or email == '' or first_name is None or \
                first_name == '' or last_name is None or last_name == '':
          continue
        user = User(first_name, last_name, email, email)  # username is email
        user_list.append(user)

  if len(application_list) > 10:
    if not click.confirm(
            f'You are adding {len(application_list)} applications to the '
            f'PoV account, continue?'):
      sys.exit(0)

  if len(user_list) > 10:
    if not click.confirm(
            f'You are adding {len(application_list)} users to the '
            f'PoV account, continue?'):
      sys.exit(0)

  add_applications_to_platform(application_list, config, setting_dict)
  add_users_to_platform(user_list, config, setting_dict)
